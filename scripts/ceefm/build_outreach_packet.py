from __future__ import annotations

import csv
import re
from collections import Counter
from datetime import date
from pathlib import Path

import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE = Path.home() / "Downloads" / "CEEFM-Prospecting-Database-Budapest.xlsx"
OUT_DIR = ROOT / "clients" / "ceefm" / "deliverables" / "outreach"
RUN_DATE = date(2026, 5, 15).isoformat()

REVIEWED_CLIENT_OVERRIDES = {
    "numa group hungary": {
        "Decision Maker Name": "Ladislav Szabo",
        "Decision Maker Role": "Country Director Hungary",
        "LinkedIn URL": "linkedin.com/in/ladislav-szabo",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "EN",
        "Next Action": "Send reviewed LinkedIn DM to Ladislav Szabo",
        "Notes": "Council reviewed. LinkedIn DM only.",
    },
    "budapest residence (arr hungary kft)": {
        "Decision Maker Name": "Gabor Balogh / Imre Tevan",
        "Decision Maker Role": "Operator / senior contact",
        "Email": "info@bpresidence.hu",
        "Phone": "+36 30 590 5327",
        "LinkedIn URL": "linkedin.com/in/tevanimre",
        "Status": "Send now",
        "Channel": "Email + LinkedIn",
        "Language": "EN",
        "Next Action": "Email Gabor, LinkedIn DM Imre if needed",
        "Notes": "Council reviewed. Multi-brand consolidation angle.",
    },
    "vagabond group": {
        "Decision Maker Name": "William Hardy / Scott Haas",
        "Decision Maker Role": "Managing Director / Owner",
        "LinkedIn URL": "linkedin.com/in/william-hardy-86518a50",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "EN",
        "Next Action": "Send reviewed LinkedIn DM to William Hardy first",
        "Notes": "Council reviewed. Use Haas as fallback.",
    },
    "walk inn apartments": {
        "Decision Maker Name": "Réka Michaletzky",
        "Decision Maker Role": "Walk Inn Apartments Budapest",
        "LinkedIn URL": "hu.linkedin.com/in/reka-michaletzky",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "HU",
        "Next Action": "Send reviewed LinkedIn DM or HU email",
        "Notes": "Council reviewed. Use one language only.",
    },
    "florin apart hotel": {
        "Decision Maker Name": "Tamás Cselőtei",
        "Decision Maker Role": "General Manager",
        "LinkedIn URL": "hu.linkedin.com/in/tamás-cselőtei-7b733076",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "HU",
        "Next Action": "Send reviewed LinkedIn DM or HU email",
        "Notes": "Council reviewed.",
    },
    "managerent kft": {
        "Decision Maker Name": "Judit Szelevényi",
        "Decision Maker Role": "Managing Director",
        "LinkedIn URL": "hu.linkedin.com/in/judit-szelevényi-08b9b1b",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "HU",
        "Next Action": "Send reviewed LinkedIn DM or HU email",
        "Notes": "Council reviewed. Property management for investors.",
    },
    "társasház-management kft": {
        "Decision Maker Name": "Gábor Gerbner",
        "Decision Maker Role": "Owner and Managing Director",
        "LinkedIn URL": "hu.linkedin.com/in/gábor-gerbner-18996775",
        "Status": "Send now",
        "Channel": "LinkedIn",
        "Language": "HU",
        "Next Action": "Send reviewed LinkedIn DM or HU email",
        "Notes": "Council reviewed. HOA and property management.",
    },
    "maverick group (mellow mood hotels)": {
        "Status": "Call first",
        "Channel": "Phone",
        "Language": "EN",
        "Next Action": "Call +36 1 413 2062 and ask for operations manager or GM",
        "Notes": "Council reviewed. Hold until named DM is confirmed.",
    },
    "budapestapartments.com (ba)": {
        "Status": "Call first",
        "Channel": "Phone",
        "Language": "EN",
        "Next Action": "Call +36 20 915 1029 to qualify FM outsourcing status and decision maker",
        "Notes": "Council reviewed. Possible embedded FM contractor setup.",
    },
    "adina apartment hotels (tfe hotels)": {
        "Status": "Cold / Q4",
        "Channel": "Research",
        "Language": "EN",
        "Next Action": "Revisit in Q4 2026 for 2027 contract cycle",
        "Notes": "Council reviewed. Group procurement sits outside Budapest.",
    },
}

WAVE1_ENRICHMENTS = {
    "bqa short rent kft": {
        "Decision Maker Name": "Shani Shamel Zada",
        "Decision Maker Role": "Owner / Directing Manager",
        "Email": "may@bqa-budapest.com",
        "Phone": "+36 70 621 2725",
        "Email Verification": "Direct - confirmed",
        "Email Source": "Wave 1 enriched sheet + BQA website team page",
        "English Confidence": "High",
        "Language": "EN",
        "Next Action": "Send bilingual email to Shani after final Hungarian QA",
        "Notes": "Wave 1 says 290 apartments and direct email confirmed.",
    },
    "continental group": {
        "Decision Maker Name": "Gabor Flesch",
        "Decision Maker Role": "Director of Operations",
        "Email": "gabor@continentalgroup.hu",
        "Email Verification": "Direct - website confirmed",
        "Email Source": "Continental Group contact page",
        "English Confidence": "High",
        "Language": "EN",
        "Status": "Research first",
        "Notes": "Verify competitor risk before pitching. Website lists operations director.",
    },
    "tower international kft": {
        "Decision Maker Name": "Tamas Piros",
        "Decision Maker Role": "Property Management Director",
        "Email": "info@towerbudapest.com",
        "Email Verification": "Generic - confirmed",
        "Email Source": "Wave 1 enriched sheet / Firecrawl",
        "English Confidence": "High",
        "Language": "EN",
        "Status": "Text only - generic inbox",
        "Notes": "Wave 1 says generic email confirmed. Direct format tamas.piros@towerbudapest.com is unverified.",
    },
    "vagabond group": {
        "LinkedIn URL": "linkedin.com/in/william-hardy-86518a50",
        "DM Verification": "Named LinkedIn from reviewed sheet",
        "Email Verification": "Generic only",
        "Email Source": "Reviewed Top 10 + company generic inbox",
        "English Confidence": "High",
        "Language": "EN",
    },
    "numa group hungary": {
        "DM Verification": "Named LinkedIn from reviewed sheet",
        "Email Verification": "No email found",
        "Email Source": "Reviewed Top 10",
        "English Confidence": "High",
        "Language": "EN",
    },
    "budapest residence (arr hungary kft)": {
        "DM Verification": "Named operator from reviewed sheet",
        "Email Verification": "Generic only",
        "Email Source": "Reviewed Top 10",
        "English Confidence": "High",
        "Language": "EN",
    },
}


CLIENT_COLUMNS = [
    "Queue ID",
    "Company Name",
    "Segment",
    "Tier",
    "Location",
    "Decision Maker",
    "Role",
    "Email",
    "Phone",
    "LinkedIn",
    "Scale",
    "Fit",
    "Pain",
    "Angle",
    "Source",
    "Status",
    "Usability Status",
    "Priority Score",
    "Channel",
    "Language",
    "English Confidence",
    "Phone Allowed",
    "Text Outreach Ready",
    "DM Verification",
    "Email Verification",
    "Email Source",
    "From Address",
    "Sender Readiness",
    "Automation Stage",
    "Automation Blocker",
    "Send Batch",
    "Send Approved",
    "Subject EN",
    "Subject HU",
    "Follow-up Due",
    "Sent At",
    "Reply Status",
    "Thread ID",
    "Next Action",
    "First Message EN",
    "First Message HU",
    "First Message",
    "Follow Up 1",
    "Notes",
]

REFERRAL_COLUMNS = [
    "Queue ID",
    "Name / Company",
    "Referral Type",
    "Location",
    "Contact Person",
    "Role",
    "Email",
    "Phone",
    "LinkedIn",
    "Who They May Know",
    "Referral Potential",
    "Status",
    "Usability Status",
    "Priority Score",
    "Channel",
    "Language",
    "English Confidence",
    "Phone Allowed",
    "Text Outreach Ready",
    "DM Verification",
    "Email Verification",
    "Email Source",
    "From Address",
    "Sender Readiness",
    "Automation Stage",
    "Automation Blocker",
    "Send Batch",
    "Send Approved",
    "Subject EN",
    "Subject HU",
    "Follow-up Due",
    "Sent At",
    "Reply Status",
    "Thread ID",
    "Next Action",
    "First Message EN",
    "First Message HU",
    "First Message",
    "Follow Up 1",
    "Commission Notes",
    "Notes",
]


def clean(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def has_value(value: str) -> bool:
    return bool(value) and value.lower() not in {"not found", "no email found"}


def has_named_person(value: str) -> bool:
    if not has_value(value):
        return False
    lowered = value.lower()
    return lowered not in {"owner", "gm", "md", "director", "operations", "property manager"}


def tier_points(tier: str) -> int:
    match = re.search(r"(\d+)", tier)
    if not match:
        return 1
    tier_num = int(match.group(1))
    return max(1, 6 - tier_num)


def language_for(text: str, location: str = "") -> str:
    blob = f"{text} {location}".lower()
    hungarian_signals = [
        "társasház",
        "ingatlan",
        "közös",
        "kft",
        "lakó",
        "hungarian",
        "buda",
        "pest",
    ]
    if any(signal in blob for signal in hungarian_signals):
        return "HU"
    return "EN"


def client_status(row: dict[str, str]) -> str:
    name = row["Decision Maker Name"]
    email = row["Email"]
    linkedin = row["LinkedIn URL"]
    notes = f"{row['Notes']} {row['Next Action']} {row['Pain Hypothesis']}".lower()

    if any(word in notes for word in ["do not pitch", "investigate first", "verify scale", "park"]):
        return "Research first"
    if has_named_person(name) and (has_value(email) or has_value(linkedin)):
        return "Send now"
    if has_value(email) or has_value(linkedin) or has_value(row["Phone"]):
        return "Qualify then send"
    return "Find contact"


def apply_reviewed_override(row: dict[str, str]) -> dict[str, str]:
    override = REVIEWED_CLIENT_OVERRIDES.get(row["Company Name"].lower(), {})
    enrichment = WAVE1_ENRICHMENTS.get(row["Company Name"].lower(), {})
    if not override and not enrichment:
        return row
    merged = row.copy()
    for key, value in {**override, **enrichment}.items():
        merged[key] = value
    return merged


def english_confidence(row: dict[str, str], lang: str) -> str:
    if row.get("English Confidence"):
        return row["English Confidence"]
    blob = f"{row.get('Company Name', '')} {row.get('Website', '')} {row.get('Segment', '')} {row.get('Why They Are a Fit', '')} {row.get('Decision Maker Role', '')}".lower()
    if any(token in blob for token in ["international", "hotel", "aparthotel", "serviced", "relocation", "foreign", "student", "luxury", "boutique"]):
        return "Medium"
    if lang == "EN":
        return "Medium"
    return "Low"


def dm_verification(row: dict[str, str], status: str) -> str:
    if row.get("DM Verification"):
        return row["DM Verification"]
    if has_named_person(row.get("Decision Maker Name", "")):
        return "Named in source workbook"
    if status in {"Call first", "Find contact"}:
        return "Missing"
    return "Role only / unverified"


def email_verification(row: dict[str, str]) -> str:
    if row.get("Email Verification"):
        return row["Email Verification"]
    email = row.get("Email", "")
    if not has_value(email):
        return "Missing"
    if email.startswith(("info@", "office@", "hotel@", "reservation@", "sales@", "hello@")):
        return "Generic only"
    return "Unverified direct-looking"


def email_source(row: dict[str, str]) -> str:
    if row.get("Email Source"):
        return row["Email Source"]
    if has_value(row.get("Email", "")):
        return "Source workbook"
    return ""


def phone_allowed(status: str, confidence: str) -> str:
    if status == "Call first":
        return "No - needs English speaker or Hungarian assist"
    if confidence == "High":
        return "Only after text reply"
    return "No"


def usability_status(status: str, lang: str, confidence: str, dm_status: str, email_status: str, channel_value: str) -> str:
    if status == "Cold / Q4":
        return "Park"
    if status == "Research first":
        return "Research before outreach"
    if dm_status == "Missing":
        return "Find verified decision maker"
    if channel_value == "Phone":
        return "Do not phone yet"
    if email_status in {"Direct - confirmed", "Direct - website confirmed"} or "LinkedIn" in channel_value:
        return "English text-ready"
    if email_status in {"Generic only", "Generic - confirmed"}:
        return "Text only - generic inbox"
    return "Needs verification"


def text_ready(usability: str) -> str:
    return "Yes" if usability in {"English text-ready", "Text only - generic inbox"} else "No"


def automation_stage(usability: str, email_status: str, channel_value: str) -> str:
    if usability == "English text-ready" and email_status in {"Direct - confirmed", "Direct - website confirmed"}:
        return "Auto-email eligible"
    if usability == "English text-ready" and "LinkedIn" in channel_value:
        return "Manual LinkedIn step"
    if usability == "Text only - generic inbox":
        return "Manual approval before email"
    if usability == "Find verified decision maker":
        return "Enrichment needed"
    return "Hold"


def automation_blocker(usability: str, dm_status: str, email_status: str) -> str:
    if usability == "English text-ready":
        return ""
    blockers = []
    if dm_status == "Missing":
        blockers.append("missing verified decision maker")
    if email_status in {"Missing", "Generic only", "Generic - confirmed", "Unverified direct-looking"}:
        blockers.append(f"email {email_status.lower()}")
    if usability == "Do not phone yet":
        blockers.append("phone call requires English speaker confirmation")
    if usability == "Research before outreach":
        blockers.append("research first")
    if usability == "Park":
        blockers.append("parked")
    return "; ".join(dict.fromkeys(blockers))


def sender_readiness(channel_value: str, automation_stage_value: str) -> str:
    if "LinkedIn" in channel_value:
        return "Not relevant - manual LinkedIn"
    if automation_stage_value == "Auto-email eligible":
        return "Blocked until Google DKIM/SPF alignment is fixed"
    if channel_value == "Email":
        return "Manual send only until sender DNS is fixed"
    return "Not ready for automated email"


def send_batch(automation_stage_value: str, priority_score: int) -> str:
    if automation_stage_value == "Auto-email eligible":
        return "Batch 1 - direct verified email"
    if automation_stage_value == "Manual LinkedIn step":
        return "Batch 1 - manual LinkedIn"
    if automation_stage_value == "Manual approval before email" and priority_score >= 12:
        return "Batch 2 - approve generic inbox"
    return "Hold"


def subject_en(row: dict[str, str]) -> str:
    company = row.get("Company Name") or row.get("Name / Company", "your properties")
    if "hotel" in f"{row.get('Segment', '')} {company}".lower() or "apartment" in f"{row.get('Segment', '')} {company}".lower():
        return f"FM support before Budapest peak season"
    if "property" in f"{row.get('Segment', '')} {company}".lower() or "társasház" in company.lower():
        return "Cleaning and maintenance support for managed properties"
    return "Budapest FM partnership"


def subject_hu(row: dict[str, str]) -> str:
    company = row.get("Company Name") or row.get("Name / Company", "")
    if "hotel" in f"{row.get('Segment', '')} {company}".lower() or "apartment" in f"{row.get('Segment', '')} {company}".lower():
        return "FM támogatás a budapesti főszezon előtt"
    if "property" in f"{row.get('Segment', '')} {company}".lower() or "társasház" in company.lower():
        return "Takarítási és karbantartási támogatás kezelt ingatlanokhoz"
    return "Budapesti FM együttműködés"


def referral_status(row: dict[str, str]) -> str:
    if has_value(row["Email"]) or has_value(row["LinkedIn URL"]):
        return "Send now"
    if has_value(row["Phone"]):
        return "Call first"
    return "Find contact"


def channel(email: str, linkedin: str, phone: str, status: str) -> str:
    if status in {"Find contact", "Research first"} and has_value(phone):
        return "Phone"
    if has_value(linkedin):
        return "LinkedIn"
    if has_value(email):
        return "Email"
    if has_value(phone):
        return "Phone"
    return "Research"


def client_score(row: dict[str, str], status: str) -> int:
    score = tier_points(row["Tier"]) * 2
    scale = row["Property Count / Estimated Scale"].lower()
    notes = f"{row['Notes']} {row['Why They Are a Fit']} {row['Pain Hypothesis']}".lower()
    if any(token in scale for token in ["400", "450", "6500", "800", "multiple", "+"]):
        score += 3
    if any(token in notes for token in ["top priority", "largest", "multi-property", "volume"]):
        score += 3
    if status == "Send now":
        score += 3
    elif status == "Qualify then send":
        score += 1
    elif status == "Research first":
        score -= 2
    return max(score, 1)


def referral_score(row: dict[str, str], status: str) -> int:
    potential = row["Referral Potential"].lower()
    score = 3
    if "high" in potential:
        score += 4
    if "medium" in potential:
        score += 2
    if status == "Send now":
        score += 2
    elif status == "Call first":
        score += 1
    return score


def client_message(row: dict[str, str], lang: str) -> str:
    status = row.get("_status", "")
    if status == "Cold / Q4":
        return "Do not send now. Revisit in Q4 2026 for the 2027 contract cycle."
    if status in {"Call first", "Find contact"}:
        return "Call or research first. Get a named operations, GM, property, or FM decision maker before sending written outreach."

    dm_value = row["Decision Maker Name"] if has_named_person(row["Decision Maker Name"]) else ""
    dm = dm_value.split("/")[0].strip().split()[0] if dm_value else "there"
    company = row["Company Name"]
    segment = row["Segment"].lower()
    pain = row["Pain Hypothesis"] or "vendor reliability becomes visible fast"
    angle = row["Suggested Outreach Angle"] or "one accountable FM partner"

    if lang == "HU":
        greeting = f"Szia {dm}" if dm_value else "Üdvözlöm"
        return (
            f"{greeting}, a(z) {company} típusú működésnél a takarítás és karbantartás "
            f"megbízhatósága gyorsan ügyfélpanasszá válik. A CEEFM budapesti FM partnerként "
            f"segít ezt egy átlátható havi szerződésbe rendezni. Összefér egy 15 perces hívás?"
        )
    if "hotel" in segment or "apartment" in segment or "hostel" in segment:
        return (
            f"Hi {dm}, {company} looks like the kind of Budapest operator where guest turnover, "
            f"cleaning reliability, and maintenance handoffs can decide review quality during peak season. "
            f"We work with CEEFM to consolidate FM under one local provider. Worth a 15-minute call?"
        )
    if "property" in segment or "hoa" in segment or "residential" in segment:
        return (
            f"Hi {dm}, managing distributed properties means owner trust depends on vendors showing up "
            f"consistently. CEEFM helps Budapest property managers replace ad hoc cleaning and maintenance "
            f"with one accountable FM contract. Worth a 15-minute call?"
        )
    return (
        f"Hi {dm}, I noticed {company} and the likely operations pressure around {pain}. "
        f"CEEFM helps Budapest operators simplify FM delivery through {angle}. Worth a quick call?"
    )


def client_follow_up(row: dict[str, str]) -> str:
    if row.get("_status", "") in {"Call first", "Find contact", "Cold / Q4"}:
        return "No written follow-up until the row has a named contact and send status."
    return (
        "Quick follow-up. The useful part may simply be a short benchmark: where cleaning, maintenance, "
        "and handoff risk usually leak time before peak season. Open to a 15-minute call?"
    )


def bilingual_message(en_message: str, hu_message: str) -> str:
    if en_message.startswith("Do not send") or en_message.startswith("Call or research"):
        return en_message
    return f"{en_message}\n\n---\n\n{hu_message}"


def referral_message(row: dict[str, str], lang: str) -> str:
    contact = row["Contact Person"] if has_named_person(row["Contact Person"]) else "there"
    org = row["Name / Company"]
    who = row["Who They May Know"] or "property operators and relocation clients"

    if lang == "HU":
        return (
            f"Szia {contact}, a CEEFM budapesti takarítási és létesítményüzemeltetési partner. "
            f"Olyan ügyfeleknek lehet hasznos, akiknél a beköltözés, vendégváltás vagy ingatlanüzemeltetés "
            f"megbízható helyi szolgáltatót igényel. Van értelme egy rövid partneri hívásnak?"
        )
    return (
        f"Hi {contact}, {org} likely sits close to {who}. CEEFM is a Budapest FM and cleaning partner "
        f"for apartments, relocation, hospitality, and managed properties. Could be a useful referral partner. "
        f"Open to a 15-minute intro?"
    )


def referral_follow_up() -> str:
    return (
        "Quick follow-up. If you ever need a reliable Budapest cleaning or FM partner for clients, "
        "CEEFM may be useful. Happy to send a short one-page partner note."
    )


def rows_from_dataframe(df: pd.DataFrame) -> list[dict[str, str]]:
    return [{column: clean(row.get(column, "")) for column in df.columns} for _, row in df.iterrows()]


def build(source: Path = DEFAULT_SOURCE) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    clients_raw = rows_from_dataframe(pd.read_excel(source, sheet_name="CEEFM Client Prospects"))
    referrals_raw = rows_from_dataframe(pd.read_excel(source, sheet_name="CEEFM Referral Partners"))

    client_rows = []
    for index, row in enumerate(clients_raw, start=1):
        row = apply_reviewed_override(row)
        override = REVIEWED_CLIENT_OVERRIDES.get(row["Company Name"].lower(), {})
        enrichment = WAVE1_ENRICHMENTS.get(row["Company Name"].lower(), {})
        status = enrichment.get("Status", override.get("Status", client_status(row)))
        row["_status"] = status
        lang = language_for(f"{row['Company Name']} {row['Segment']} {row['Decision Maker Role']}", row["District / Location"])
        base_lang = enrichment.get("Language", override.get("Language", lang))
        lang = "EN+HU"
        selected_channel = override.get("Channel", channel(row["Email"], row["LinkedIn URL"], row["Phone"], status))
        confidence = english_confidence(row, base_lang)
        dm_status = dm_verification(row, status)
        mail_status = email_verification(row)
        usability = usability_status(status, base_lang, confidence, dm_status, mail_status, selected_channel)
        stage = automation_stage(usability, mail_status, selected_channel)
        score = client_score(row, status)
        message_en = client_message(row, "EN")
        message_hu = client_message(row, "HU")
        client_rows.append(
            {
                "Queue ID": f"CEEFM-C-{index:03d}",
                "Company Name": row["Company Name"],
                "Segment": row["Segment"],
                "Tier": row["Tier"],
                "Location": row["District / Location"],
                "Decision Maker": row["Decision Maker Name"],
                "Role": row["Decision Maker Role"],
                "Email": row["Email"],
                "Phone": row["Phone"],
                "LinkedIn": row["LinkedIn URL"],
                "Scale": row["Property Count / Estimated Scale"],
                "Fit": row["Why They Are a Fit"],
                "Pain": row["Pain Hypothesis"],
                "Angle": row["Suggested Outreach Angle"],
                "Source": row["Source URL"],
                "Status": status,
                "Usability Status": usability,
                "Priority Score": score,
                "Channel": selected_channel,
                "Language": lang,
                "English Confidence": confidence,
                "Phone Allowed": phone_allowed(status, confidence),
                "Text Outreach Ready": text_ready(usability),
                "DM Verification": dm_status,
                "Email Verification": mail_status,
                "Email Source": email_source(row),
                "From Address": "office.ceefm@gmail.com via office@ceefm.eu forwarder setup",
                "Sender Readiness": sender_readiness(selected_channel, stage),
                "Automation Stage": stage,
                "Automation Blocker": automation_blocker(usability, dm_status, mail_status),
                "Send Batch": send_batch(stage, score),
                "Send Approved": "No",
                "Subject EN": subject_en(row),
                "Subject HU": subject_hu(row),
                "Follow-up Due": "",
                "Sent At": "",
                "Reply Status": "",
                "Thread ID": "",
                "Next Action": override.get("Next Action", row["Next Action"]),
                "First Message EN": message_en,
                "First Message HU": message_hu,
                "First Message": bilingual_message(message_en, message_hu),
                "Follow Up 1": client_follow_up(row),
                "Notes": override.get("Notes", row["Notes"]),
            }
        )

    referral_rows = []
    for index, row in enumerate(referrals_raw, start=1):
        status = referral_status(row)
        base_lang = language_for(f"{row['Name / Company']} {row['Referral Type']} {row['Role']}", row["Location / District"])
        lang = "EN+HU"
        selected_channel = channel(row["Email"], row["LinkedIn URL"], row["Phone"], status)
        confidence = english_confidence(
            {
                "Company Name": row["Name / Company"],
                "Website": row["Website"],
                "Segment": row["Referral Type"],
                "Why They Are a Fit": row["Who They May Know"],
                "Decision Maker Role": row["Role"],
                "English Confidence": "",
            },
            base_lang,
        )
        dm_status = "Named in source workbook" if has_named_person(row["Contact Person"]) else "Missing"
        mail_status = email_verification({"Email": row["Email"], "Email Verification": ""})
        usability = usability_status(status, base_lang, confidence, dm_status, mail_status, selected_channel)
        stage = automation_stage(usability, mail_status, selected_channel)
        score = referral_score(row, status)
        message_en = referral_message(row, "EN")
        message_hu = referral_message(row, "HU")
        referral_rows.append(
            {
                "Queue ID": f"CEEFM-R-{index:03d}",
                "Name / Company": row["Name / Company"],
                "Referral Type": row["Referral Type"],
                "Location": row["Location / District"],
                "Contact Person": row["Contact Person"],
                "Role": row["Role"],
                "Email": row["Email"],
                "Phone": row["Phone"],
                "LinkedIn": row["LinkedIn URL"],
                "Who They May Know": row["Who They May Know"],
                "Referral Potential": row["Referral Potential"],
                "Status": status,
                "Usability Status": usability,
                "Priority Score": score,
                "Channel": selected_channel,
                "Language": lang,
                "English Confidence": confidence,
                "Phone Allowed": phone_allowed(status, confidence),
                "Text Outreach Ready": text_ready(usability),
                "DM Verification": dm_status,
                "Email Verification": mail_status,
                "Email Source": "Source workbook" if has_value(row["Email"]) else "",
                "From Address": "office.ceefm@gmail.com via office@ceefm.eu forwarder setup",
                "Sender Readiness": sender_readiness(selected_channel, stage),
                "Automation Stage": stage,
                "Automation Blocker": automation_blocker(usability, dm_status, mail_status),
                "Send Batch": send_batch(stage, score),
                "Send Approved": "No",
                "Subject EN": "CEEFM referral partnership in Budapest",
                "Subject HU": "CEEFM partneri együttműködés Budapesten",
                "Follow-up Due": "",
                "Sent At": "",
                "Reply Status": "",
                "Thread ID": "",
                "Next Action": row["Next Action"],
                "First Message EN": message_en,
                "First Message HU": message_hu,
                "First Message": bilingual_message(message_en, message_hu),
                "Follow Up 1": referral_follow_up(),
                "Commission Notes": row["Commission Split Notes"],
                "Notes": row["Notes"],
            }
        )

    client_rows.sort(key=lambda r: (-int(r["Priority Score"]), r["Status"], r["Company Name"]))
    referral_rows.sort(key=lambda r: (-int(r["Priority Score"]), r["Status"], r["Name / Company"]))

    client_csv = OUT_DIR / f"CEEFM-client-outreach-queue-{RUN_DATE}.csv"
    referral_csv = OUT_DIR / f"CEEFM-referral-partner-queue-{RUN_DATE}.csv"
    with client_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CLIENT_COLUMNS)
        writer.writeheader()
        writer.writerows(client_rows)
    with referral_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REFERRAL_COLUMNS)
        writer.writeheader()
        writer.writerows(referral_rows)

    workbook_path = OUT_DIR / f"CEEFM-full-outreach-packet-{RUN_DATE}.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(client_rows, columns=CLIENT_COLUMNS).to_excel(writer, index=False, sheet_name="Client Queue")
        text_ready_clients = [r for r in client_rows if r["Text Outreach Ready"] == "Yes"]
        pd.DataFrame(text_ready_clients, columns=CLIENT_COLUMNS).to_excel(writer, index=False, sheet_name="Text Ready Queue")
        send_control_clients = [r for r in client_rows if r["Send Batch"] != "Hold"]
        pd.DataFrame(send_control_clients, columns=CLIENT_COLUMNS).to_excel(writer, index=False, sheet_name="Send Control")
        missing_clients = [r for r in client_rows if r["Text Outreach Ready"] != "Yes" or r["Automation Blocker"]]
        pd.DataFrame(missing_clients, columns=CLIENT_COLUMNS).to_excel(writer, index=False, sheet_name="Missing Fields")
        pd.DataFrame(referral_rows, columns=REFERRAL_COLUMNS).to_excel(writer, index=False, sheet_name="Referral Queue")
        pd.DataFrame(summary_rows(client_rows, referral_rows)).to_excel(writer, index=False, sheet_name="Dashboard")
        pd.DataFrame(automation_rows()).to_excel(writer, index=False, sheet_name="Automation Plan")
        pd.DataFrame(playbook_rows()).to_excel(writer, index=False, sheet_name="Scripts")

    format_workbook(workbook_path)

    playbook = OUT_DIR / f"CEEFM-full-outreach-playbook-{RUN_DATE}.md"
    playbook.write_text(render_markdown(client_rows, referral_rows), encoding="utf-8")


def summary_rows(client_rows: list[dict[str, str]], referral_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    rows.append({"Metric": "Client prospects", "Value": len(client_rows), "Notes": "Full database, not Top 10 only"})
    rows.append({"Metric": "Referral partners", "Value": len(referral_rows), "Notes": "Partner and channel sources"})
    rows.append({"Metric": "Client text-ready rows", "Value": sum(r["Text Outreach Ready"] == "Yes" for r in client_rows), "Notes": "Bilingual written outreach allowed"})
    rows.append({"Metric": "Client auto-email eligible", "Value": sum(r["Automation Stage"] == "Auto-email eligible" for r in client_rows), "Notes": "Direct verified email only"})
    rows.append({"Metric": "Client manual LinkedIn rows", "Value": sum(r["Automation Stage"] == "Manual LinkedIn step" for r in client_rows), "Notes": "LinkedIn should stay manual"})
    for status, count in Counter(r["Status"] for r in client_rows).items():
        rows.append({"Metric": f"Client status: {status}", "Value": count, "Notes": ""})
    for status, count in Counter(r["Status"] for r in referral_rows).items():
        rows.append({"Metric": f"Referral status: {status}", "Value": count, "Notes": ""})
    rows.append({"Metric": "Total immediate sends", "Value": sum(r["Status"] == "Send now" for r in client_rows + referral_rows), "Notes": "Still needs human review before send"})
    return rows


def automation_rows() -> list[dict[str, str]]:
    return [
        {
            "Layer": "Source of truth",
            "Rule": "Use Text Ready Queue for daily execution. Do not automate from raw Client Queue.",
            "Tool": "Google Sheets or local XLSX",
            "Status": "Ready",
        },
        {
            "Layer": "Sender identity",
            "Rule": "Send from office.ceefm@gmail.com configured to appear as office@ceefm.eu only after DNS alignment is fixed. Current check: SPF authorizes Hostinger, DMARC exists with p=none, Google DKIM selector is missing.",
            "Tool": "Gmail / Google Workspace DNS",
            "Status": "Blocked",
        },
        {
            "Layer": "Email automation",
            "Rule": "Only auto-send rows with Automation Stage = Auto-email eligible and Send Approved = Yes. Keep sending paused until Gmail-as-domain DNS is corrected.",
            "Tool": "Gmail API, Apps Script, n8n, or Make",
            "Status": "Built but paused",
        },
        {
            "Layer": "LinkedIn",
            "Rule": "Do not automate LinkedIn sending. Generate copy, then manually send connection or DM.",
            "Tool": "Manual LinkedIn",
            "Status": "Manual",
        },
        {
            "Layer": "Phone",
            "Rule": "No cold phone calls unless English speaker is verified or a Hungarian-speaking helper is available.",
            "Tool": "Manual",
            "Status": "Gated",
        },
        {
            "Layer": "Follow-up",
            "Rule": "Create follow-up tasks 3 business days after first touch. Stop when reply received.",
            "Tool": "Google Sheets + Calendar/Gmail task",
            "Status": "Ready",
        },
    ]


def playbook_rows() -> list[dict[str, str]]:
    return [
        {
            "Script": "Call for missing decision maker",
            "Copy": "Hi, I am trying to reach the person responsible for cleaning, maintenance, or FM supplier decisions. Who is the best person to speak with?",
        },
        {
            "Script": "LinkedIn connection note",
            "Copy": "Hi {{FirstName}}, I work with CEEFM on Budapest facility management and cleaning partnerships. Thought it may be relevant to your properties.",
        },
        {
            "Script": "Reply if interested",
            "Copy": "Great. The simplest next step is a 15-minute call to check current vendor setup, property count, and whether there is a 2026 need. I can send two time options.",
        },
        {
            "Script": "Referral ask",
            "Copy": "If you know one Budapest property or relocation operator who needs reliable cleaning or FM support, an introduction would be useful. We can keep it simple and professional.",
        },
        {
            "Script": "Quality rule",
            "Copy": "Do not send to a generic inbox without a named person unless the row is marked Qualify then send and the message references a specific property or operational signal.",
        },
    ]


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="0F1A2E")
    header_font = Font(color="F5F0E8", bold=True)
    accent_fill = PatternFill("solid", fgColor="D4A843")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "A": 14,
            "B": 30,
            "C": 28,
            "D": 14,
            "E": 22,
            "F": 22,
            "G": 22,
            "H": 26,
            "I": 18,
            "J": 28,
            "K": 26,
            "L": 38,
            "M": 34,
            "N": 34,
            "O": 28,
            "P": 18,
            "Q": 14,
            "R": 16,
            "S": 12,
            "T": 36,
            "U": 68,
            "V": 52,
            "W": 40,
        }
        for column_index in range(1, ws.max_column + 1):
            column = get_column_letter(column_index)
            ws.column_dimensions[column].width = widths.get(column, 22)
        for idx in range(2, ws.max_row + 1):
            ws.row_dimensions[idx].height = 54
        if ws.title == "Dashboard":
            ws["A1"].fill = accent_fill
            ws["B1"].fill = accent_fill
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 48

    wb.save(path)


def markdown_table(rows: list[dict[str, str]], fields: list[str], limit: int = 20) -> str:
    selected = rows[:limit]
    header = "| " + " | ".join(fields) + " |"
    sep = "| " + " | ".join(["---"] * len(fields)) + " |"
    body = []
    for row in selected:
        body.append("| " + " | ".join(str(row.get(field, "")).replace("|", "/") for field in fields) + " |")
    return "\n".join([header, sep, *body])


def render_markdown(client_rows: list[dict[str, str]], referral_rows: list[dict[str, str]]) -> str:
    client_counts = Counter(r["Status"] for r in client_rows)
    referral_counts = Counter(r["Status"] for r in referral_rows)
    send_clients = [r for r in client_rows if r["Status"] == "Send now"]
    send_refs = [r for r in referral_rows if r["Status"] == "Send now"]
    hold_clients = [r for r in client_rows if r["Status"] != "Send now"]

    return f"""# CEEFM Full Outreach Packet

**Built:** {RUN_DATE}
**Scope:** 95 client prospects plus 51 referral partners.
**Source:** CEEFM-Prospecting-Database-Budapest.xlsx.

## Executive Summary

- Client prospects: {len(client_rows)}
- Referral partners: {len(referral_rows)}
- Client send now: {client_counts.get("Send now", 0)}
- Client qualify then send: {client_counts.get("Qualify then send", 0)}
- Client research first: {client_counts.get("Research first", 0)}
- Client find contact: {client_counts.get("Find contact", 0)}
- Referral send now: {referral_counts.get("Send now", 0)}
- Referral call first: {referral_counts.get("Call first", 0)}
- Referral find contact: {referral_counts.get("Find contact", 0)}

## Operating Rules

1. Send one language per prospect. Use HU for Hungarian-owned property management and HOA rows. Use EN for international hospitality rows.
2. Do not send a generic note to info@ when no decision maker is named.
3. If a row says Research first, complete that research before outreach. This applies especially to possible competitors and ambiguous group-procurement targets.
4. Use LinkedIn first when a named decision maker is present and a LinkedIn URL exists.
5. Use email when a direct email exists or when the row has a named person plus a credible generic company inbox.

## First Client Wave

{markdown_table(send_clients, ["Queue ID", "Company Name", "Tier", "Decision Maker", "Channel", "Language", "Priority Score"], 25)}

## First Referral Wave

{markdown_table(send_refs, ["Queue ID", "Name / Company", "Referral Type", "Channel", "Language", "Priority Score"], 20)}

## Hold And Research Queue

{markdown_table(hold_clients, ["Queue ID", "Company Name", "Status", "Channel", "Next Action"], 35)}

## Call Script

Hi, I am trying to reach the person responsible for cleaning, maintenance, or FM supplier decisions. Who is the best person to speak with?

If they ask why: CEEFM is a Budapest facility management and cleaning provider. We are checking whether there is a need for a reliable local FM partner before peak season.

## LinkedIn Connection Note

Hi {{FirstName}}, I work with CEEFM on Budapest facility management and cleaning partnerships. Thought it may be relevant to your properties.

## Send QA Checklist

- Named person present, or call-first status assigned.
- One language only.
- No mixed English inside Hungarian copy.
- Message references property type, scale, or operational pain.
- Follow-up date set 3 business days after first touch.
"""


if __name__ == "__main__":
    build()
